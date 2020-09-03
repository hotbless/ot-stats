import openpyxl


class ReadOtData:
    def __init__(self):
        self.filename = "otrec.xlsx"
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.save_name = "new_ot.xlsx"

    def unmerge(self):
        for ws in self.wb:
            # print(ws.title)
            merged_ranges = ws.merged_cells.ranges
            o = 0
            while merged_ranges:
                for entry in merged_ranges:
                    o = +1
                    print("  unMerging: " + str(o) + ": " + str(entry))
                    ws.unmerge_cells(str(entry))
        self.wb.save(self.save_name)

    def read_data(self):
        self.unmerge()
        wb = openpyxl.load_workbook(filename=self.save_name)
        return wb


class ReadWtData:
    def __init__(self):
        self.filename = "wtrec.xlsx"
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.save_name = "new_wt.xlsx"

    def unmerge(self):
        for ws in self.wb:
            # print(ws.title)
            merged_ranges = ws.merged_cells.ranges
            o = 0
            while merged_ranges:
                for entry in merged_ranges:
                    o = +1
                    print("  unMerging: " + str(o) + ": " + str(entry))
                    ws.unmerge_cells(str(entry))
        self.wb.save(self.save_name)

    def read_data(self):
        self.unmerge()
        wb = openpyxl.load_workbook(filename=self.save_name)
        return wb