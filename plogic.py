# from openpyxl.cell.cell import Cell
# from openpyxl.styles import Font
import openpyxl


from approve import OtApprv
from timem import ElasTime
from exceldata import ResToXlsx

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

from xldata import ReadOtData, ReadWtData


class LogicStr:
    def __init__(self, logic):
        self.logic = logic

    @property
    def res(self):
        restr = ''
        if self.logic == False:
            restr = '拒绝'
        else:
            if self.logic == True:
                restr = '批准'
        return restr


class CalWt:
    def __init__(self):
        self.wb = openpyxl.load_workbook(filename='new_wt.xlsx')
        self.apv_flag = True
        self.ref_cmts = []

    def match_records(self, name, date):
        # print(name)
        # print(date)
        wt_records = self.wb
        for wt_ws in wt_records:
            if wt_ws.title == "每日统计":
                wt_start = None
                wt_end = None
                wt_es = None
                for rec in wt_ws.rows:
                    wt_name = rec[0].value
                    wt_date = rec[6].value
                    if (wt_name is not None) and (wt_date is not None):
                        if (name in wt_name) and (date in wt_date):
                            wt_start = rec[8].value
                            wt_end = rec[10].value
                            if (wt_start is not None) and (wt_end is not None):
                                wt_es = ElasTime(wt_start, wt_end).elas_hrs()
                                self.apv_flag = True
                                self.ref_cmts = list()
                                break
                            else:
                                self.apv_flag = False
                                self.ref_cmts = list('工作表中无对应记录')
                        else:
                            self.apv_flag = False
                            self.ref_cmts = list('工作表中无对应记录')
                return wt_start, wt_end, wt_es, self.apv_flag, self.ref_cmts



                    # if (wt_name is not None) and (wt_date is not None):
                        # print("judge")
                        # print(name)
                        # print(wt_name)
                        # print(date)
                        # print(wt_date)
                        # if (name in wt_name) and (date in wt_date):
                        #     wt_start = rec[8].value
                        #     wt_end = rec[10].value
                        #     wt_es = ElasTime(wt_start, wt_end).elas_hrs()
                        #     return wt_start, wt_end, wt_es


class CalOt:
    def __init__(self):
        self.wb_ot = openpyxl.load_workbook(filename='new_ot.xlsx')
        self.wb_re = openpyxl.load_workbook(filename='results.xlsx')
        self.re_data = list()

    def calculate_date(self):
        ot_records = self.wb_ot
        print(ot_records)

        for ot_ws in ot_records:
            print(ot_ws)
            for rec in ot_ws.rows:
                if rec[0].value != "审批编号":
                    ot_name = rec[14].value
                    ot_start = rec[15].value
                    ot_end = rec[16].value
                    ot_elapsed = rec[20].value
                    ot_start_fmt = ot_start.split()[0][2:]
                    ot_start_time = ot_start.split()[1]
                    ot_end_fmt = ot_end.split()[0][2:]
                    ot_end_time = ot_end.split()[1]

                    wt_rec = CalWt().match_records(ot_name, ot_start_fmt)
                    # ot_is_wd = IsWorkDay(ot_start).res
                    # ot_elas_hrs = ElasTime(ot_start_time, ot_end_time).elas_hrs()
                    if wt_rec[3] is True:
                        ot_appr = OtApprv(ot_day=ot_start, ot_s=ot_start_time, ot_e=ot_end_time, wt_s=wt_rec[0], wt_e=wt_rec[1], wt_elas=wt_rec[2]).ot_res
                        print(ot_name + ' 开始于 ' +ot_start + ' 的 ' + ot_elapsed + ' 小时加班审核结果是: '+str(ot_appr[0])+''.join(ot_appr[1]))
                        self.re_data.append([ot_name, ot_start, ot_end, ot_elapsed, LogicStr(ot_appr[0]).res, ''.join(ot_appr[1])])
                        print(self.re_data)
                    else:
                        print(str(wt_rec[3]) + ''.join(wt_rec[4]))
                        self.re_data.append([ot_name, ot_start, ot_end, ot_elapsed, LogicStr(wt_rec[3]).res, ''.join(wt_rec[4])])
                        print(self.re_data)

        ws_re = self.wb_re['批准结果']
        for row in self.re_data:
            for each in row:
                if each == '批准':
                    each = openpyxl.cell.Cell(ws_re)
                    each.font = openpyxl.styles.Font(bold=True, color="00FF00")
                else:
                    if each == '拒绝':
                        each = openpyxl.cell.Cell(ws_re)
                        each.font = openpyxl.styles.Font(bold=True, color="FF0000")
                # yield each
            ws_re.append(row)
        self.wb_re.save(filename='results.xlsx')
        for cl in ws_re['E']:
            if cl.value == '批准':
                cl.font = openpyxl.styles.Font(bold=True, color="00FF00")
            else:
                if cl.value == '拒绝':
                    cl.font = openpyxl.styles.Font(bold=True, color="FF0000")
        self.wb_re.save(filename='results.xlsx')
                    # print(ot_is_wd)
                    # print(wt_rec)


if __name__ == '__main__':
    ReadOtData().read_data()
    ReadWtData().read_data()
    ResToXlsx().res_to_file()
    CalOt().calculate_date()
