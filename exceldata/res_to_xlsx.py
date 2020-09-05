import openpyxl
import os


class ResToXlsx:
    def __init__(self):
        self.res_filename = 'results.xlsx'

    def xlsx_header(self, ws):
        ws['A1'] = '姓名'
        ws['B1'] = '加班开始'
        ws['C1'] = '加班结束'
        ws['D1'] = '申请时长'
        ws['E1'] = '计算结果'
        ws['F1'] = '备注'

    def res_to_file(self):
        if os.path.exists(self.res_filename):
            os.remove(self.res_filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '批准结果'
        self.xlsx_header(ws)
        wb.save(filename=self.res_filename)