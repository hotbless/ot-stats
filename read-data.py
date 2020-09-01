import openpyxl
import datetime
import chinese_calendar



# class CalDatetime:
#     def __init__(self, str_time):
#         self.dt_day = str_time.split()[0]
#         self.dt_time = str_time.split()[1]
#
#     def dt_hrmin(self):
#         self.dt_time = datetime.datetime.strptime(self.dt_time, '%H:%M')
#
#     def dt_ymd(self):
#         self.dt_day = datetime.datetime.strptime(self.dt_day, '%Y-%m-%d')
#
#     @property
#     def res(self):
#         self.dt_ymd()
#         self.dt_hrmin()
#         return self.dt_day, self.dt_time


class CalDatetime:
    def __init__(self, str_time):
        self.time_tr = str_time

    def dt_hrmin(self):
        hm = datetime.datetime.strptime(self.time_tr, '%H:%M')
        return hm

    def dt_ymd(self):
        ymd = datetime.datetime.strptime(self.time_tr, '%Y-%m-%d')
        return ymd

    def dt_all(self):
        all = datetime.datetime.strptime(self.time_tr, '%Y-%m-%d %H:%M')
        return all



class ElasTime:
    def __init__(self, time_start, time_end):
        self.time_start = time_start
        self.time_end = time_end

    def hm_str_to_dt(self):
        self.time_start = CalDatetime(self.time_start).dt_hrmin()
        self.time_end = CalDatetime(self.time_end).dt_hrmin()

    def elas_hrs(self):
        self.hm_str_to_dt()
        es = (self.time_end - self.time_start).seconds/3600
        return es


class IsWorkDay:
    def __init__(self, w_date):
        # w_date_f = datetime.datetime.strptime(w_date, '%Y-%m-%d %H:%M')

        # self.w_res = chinese_calendar.is_workday(w_date_f)
        w_date_f = CalDatetime(w_date).dt_all()
        self.w_res = chinese_calendar.is_workday(w_date_f)
        on_holiday, holiday_name = chinese_calendar.get_holiday_detail(w_date_f)
        print("is work day?", self.w_res)
        print("is holiday?", on_holiday)
        print("holiday name?", holiday_name)

    @property
    def res(self):
        return self.w_res


class ReadOtData:
    def __init__(self):
        self.filename = "20200717.xlsx"
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.save_name = "new_ot.xlsx"

    def unmerge(self):
        for ws in self.wb:
            print(ws.title)
            merged_ranges = ws.merged_cells.ranges
            o = 0
            while merged_ranges:
                for entry in merged_ranges:
                    o = +1
                    print("  unMerging: " + str(o) + ": " + str(entry))
                    ws.unmerge_cells(str(entry))
        self.wb.save(self.save_name)

    def read_data(self):
        self.unmerge()
        wb = openpyxl.load_workbook(filename=self.save_name)
        return wb


class ReadWtData:
    def __init__(self):
        self.filename = "20200630.xlsx"
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.save_name = "new_wt.xlsx"

    def unmerge(self):
        for ws in self.wb:
            print(ws.title)
            merged_ranges = ws.merged_cells.ranges
            o = 0
            while merged_ranges:
                for entry in merged_ranges:
                    o = +1
                    print("  unMerging: " + str(o) + ": " + str(entry))
                    ws.unmerge_cells(str(entry))
        self.wb.save(self.save_name)

    def read_data(self):
        self.unmerge()
        wb = openpyxl.load_workbook(filename=self.save_name)
        return wb


class CalDate:
    def __init__(self):
        pass

    def calculate_date(self):
        ot_records = ReadOtData().read_data()
        # ot_name = ot_records[0]['over']
        # print(ot_name)
        ot_start = CalDatetime(ot_records[0]['开始时间']).res
        print(ot_start)
        d_res = IsWorkDay(ot_start).res
        if d_res is True:
            print("workday")
        else:
            print("holiday")
        ot_end = CalDatetime(ot_records[0]['结束时间']).res
        print(ot_end)
        ot_mins = ElasTime(ot_start, ot_end).res / 60
        print(ot_mins)



# CalDate().calculate_date()
#     # ReadOtData().read_xlsx()
# x = ReadWtData().read_sheet()
# print(x)


class CalWt:
    def __init__(self):
        self.wb = openpyxl.load_workbook(filename='new_wt.xlsx')

    def match_records(self, name, date):
        # print(name)
        # print(date)
        wt_records = self.wb
        for wt_ws in wt_records:
            if wt_ws.title == "每日统计":
                for rec in wt_ws.rows:
                    wt_name = rec[0].value
                    wt_date = rec[6].value
                    if (wt_name is not None) and (wt_date is not None):
                        # print("judge")
                        # print(name)
                        # print(wt_name)
                        # print(date)
                        # print(wt_date)
                        if (name in wt_name) and (date in wt_date):
                            wt_start = rec[8].value
                            wt_end = rec[10].value
                            wt_es = ElasTime(wt_start, wt_end).elas_hrs()
                            return wt_start, wt_end, wt_es


class OtApprv:
    def __init__(self, ot_day, ot_start, ot_end, wt_start, wt_end, wt_elas):
        self.dt_std_edn = CalDatetime('17:30').dt_hrmin()
        self.res = False
        ot_elas_hrs = ElasTime(ot_start, ot_end).elas_hrs()
        ot_is_wd = IsWorkDay(ot_day).res
        if (ot_elas_hrs >= 2.5) and (ot_is_wd is False) and (ot_start >= wt_start) and (ot_end >= wt_end):
            self.res = True
        else:
            if (wt_elas >= 11.5) and (ot_start >= self.dt_std_edn) and (ot_start >= (wt_start + 9)) and (ot_end >= wt_end):
                self.res = True

    @property
    def ot_res(self):
        return self.res


                        

class CalOt:
    def __init__(self):
        self.wb = openpyxl.load_workbook(filename='new_ot.xlsx')

    def calculate_date(self):
        ot_records = self.wb

        for ot_ws in ot_records:
            for rec in ot_ws.rows:
                if (rec[0].value) != "审批编号":
                    ot_name = rec[14].value
                    ot_start = rec[15].value
                    ot_end = rec[16].value
                    ot_elapsed = rec[20].value
                    ot_start_fmt = ot_start.split()[0][2:]
                    ot_start_time = ot_start.split()[1]
                    ot_end_fmt = ot_end.split()[0][2:]
                    ot_end_time = ot_end.split()[1]

                    wt_rec = CalWt().match_records(ot_name, ot_start_fmt)
                    ot_is_wd = IsWorkDay(ot_start).res
                    # ot_elas_hrs = ElasTime(ot_start_time, ot_end_time).elas_hrs()

                    ot_appr = OtApprv(otday=ot_start, ot_start=ot_start_time, ot_end=ot_end_time, wt_start=wt_rec[0], wt_end=wt_rec[1], wt_elas=wt_rec[2]).res
                    print(ot_appr)



                    print(ot_is_wd)
                    print(wt_rec)
                    # x = CalDatetime(ot_start).res
                    # print(x)
                    #
                    # print(ot_name)
                    # print(ot_start)
                    # print(ot_end)
                    # print(ot_elapsed)






        # print(ot_records.worksheets[0].title)
        # ot_start = ot_records.worksheets[0]['P2'].value
        # print(ot_start)
        # ots = ot_records.worksheets[0]['P']
        # for i in ots:
        #     # if i is not ots[0]:
        #     if i.value is not '开始时间':
        #         print(i.value)
        # print(type(ot_start))
        # ot_start = StrToDatetimeHM(ot_records.worksheets[0]['p2'].value).res
        # print(ot_start)

    def unmerge(self, wb, save_name):
        for ws in wb:
            print(ws.title)
            merged_ranges = ws.merged_cells.ranges
            o = 0
            while merged_ranges:
                for entry in merged_ranges:
                    o = +1
                    print("  unMerging: " + str(o) + ": " + str(entry))
                    ws.unmerge_cells(str(entry))
        wb.save(save_name)
        
        
        
        
class XTest:
    def __init__(self):
        pass


    def dtcompr(self):
        std_end = '2020-06-02 17:30'
        wt_end = '2020-06-03 17:30'
        dt_std_end = CalDatetime(std_end).dt_all()
        dt_wt_end = CalDatetime(wt_end).dt_all()
        if dt_wt_end >= dt_std_end:
            print('True')
        else:
            print('False')
        
        
        
        
# XTest().dtcompr()

# wt_wb = ReadWtData().read_data()
# ot_wb = ReadOtData().read_data()
CalOt().calculate_date()
